import streamlit as st
import pandas as pd
import openpyxl
import matplotlib.pyplot as plt
from openpyxl.drawing.image import Image
import numpy as np
import io
import xlsxwriter
import math
from openpyxl.styles import Border, Side
from openpyxl.styles import PatternFill
from openpyxl.chart import LineChart, Reference
import xlsxwriter
import pandas as pd
import matplotlib.pyplot as plt
from io import BytesIO
import openpyxl
from openpyxl.chart import LineChart
from openpyxl.drawing.colors import ColorChoice
from openpyxl.drawing.text import CharacterProperties
from openpyxl.styles import Font
from openpyxl import Workbook, load_workbook
from datetime import datetime
from openpyxl.styles import Alignment
from openpyxl.styles import PatternFill

def extract_first_digit(number):
    """ดึงเลขหลักแรกจากตัวเลข"""
    while number >= 10:
        number //= 10
    return number

# เปิดหน้าต่างให้เลือกไฟล์ GL
root = tk.Tk()
root.withdraw()  # ซ่อนหน้าต่างหลัก

if not gl_file:
    print("❌ คุณไม่ได้เลือกไฟล์ฐานข้อมูล(GL)!")
    exit()

# อ่านข้อมูลจากไฟล์ GL
print("📖 กำลังอ่านข้อมูลจากไฟล์...")

xls = pd.ExcelFile(gl_file)
gl_sheets = [s for s in xls.sheet_names if 'GL' in s]

if not gl_sheets:
    print("❌ ไม่พบชีทที่มี 'GL' ในชื่อ!")
    exit()

# ใช้ชีทแรกที่พบเป็นแหล่งข้อมูล
sheet_name = gl_sheets[0]
df_gl = pd.read_excel(gl_file, sheet_name=sheet_name)

print(f"✅ อ่านข้อมูลไฟล์ GL แล้ว")

#  ให้ผู้ใช้เลือกที่เก็บไฟล์ใหม่
if not output_filename:
    print("❌ คุณไม่ได้เลือกที่เก็บไฟล์!")
    exit()

# บันทึกข้อมูล GL ลงไฟล์ใหม่
print("⌛ กำลังสร้างไฟล์ใหม่...")
with pd.ExcelWriter(output_filename, engine="openpyxl") as writer:
    df_gl.to_excel(writer, sheet_name="GL", index=False)

print(f"📂 ไฟล์ถูกบันทึกแล้ว ที่: {output_filename}")

# ฟังก์ชันเลือกบัญชีจากลิสต์
def choose_account(matches):
    while True:
        print("🔹 พบข้อมูลที่คล้ายกัน:")
        for i, acc in enumerate(matches):
            print(f"[{i+1}] {acc}")
        print("[0] ยกเลิกการค้นหา")
        choice = input("เลือกหมายเลขบัญชีที่ต้องการ: ")
        if choice.isdigit():
            choice = int(choice)
            if choice == 0:
                print("❌ ยกเลิกการค้นหา")
                return None
            if 1 <= choice <= len(matches):
                selected_account = matches[choice - 1]  # เลือกบัญชีตามหมายเลข
                print(f"✅ คุณเลือกบัญชี: {selected_account}")  # แจ้งผลการเลือก
                return selected_account  # คืนค่าบัญชีที่เลือก
        print("⚠️ กรุณาเลือกหมายเลขที่ถูกต้อง!")
        # กรองข้อมูลตามบัญชีที่เลือก
        filtered_df = df_gl[df_gl[account_cols].apply(lambda x: any(selected_account.lower() in str(val).lower() for val in x), axis=1)]
                
        # ตรวจสอบผลลัพธ์
        if not filtered_df.empty:
            print(f"✅ พบข้อมูลที่ตรงกับบัญชีที่เลือก: {selected_account}")
            # ทำงานต่อ (เช่น สร้างกราฟ, บันทึกข้อมูล)
            # คุณสามารถเรียกใช้โค้ดที่เหลือต่อไปได้ที่นี่
            return filtered_df  # ส่งกลับข้อมูลที่กรองแล้ว
        else:
            print(f"❌ ไม่พบข้อมูลที่ตรงกับบัญชี {selected_account}")
            return None  # ไม่มีข้อมูลตรงกับบัญชีที่เลือก
        print("⚠️ กรุณาเลือกหมายเลขที่ถูกต้อง!")

# รับค่า "เลขที่บัญชี" หรือ "ชื่อบัญชี" จากผู้ใช้
while True:
    user_input = input("🔍 เลขที่บัญชีหรือชื่อบัญชีที่ต้องการตรวจสอบ: ").strip()
    
    #  ค้นหาคอลัมน์บัญชี
    account_cols = [col for col in df_gl.columns if any(keyword in col.lower() for keyword in ["เลขที่บัญชี", "account no.", "ชื่อบัญชี", "account name"])]
    if not account_cols:
        print("❌ ไม่พบคอลัมน์ที่เกี่ยวข้องกับบัญชี!")
        exit()

    # ทำความสะอาดข้อมูล (ลบช่องว่าง & ค่า NaN)
    df_gl[account_cols] = df_gl[account_cols].apply(lambda x: x.astype(str).str.strip().fillna(''))
 
    # ค้นหาข้อมูล
    matches = set()
    
    filtered_df_list = []  # กำหนดให้เป็นลิสต์ว่างก่อนใช้งาน

    for col in account_cols:
        matches.update(df_gl[col].dropna().astype(str).unique())
    
    matches = sorted([m for m in matches if user_input.lower() in str(m).lower()])
           
    if not matches:
        print("⚠️ กรุณาลองใหม่!")
        continue
    
    if len(matches) > 5:  # If there are too many matches, ask for a more specific query
        print("⚠️ กรุณากรอกคำค้นหาใหม่!")
        continue
    
    if len(matches) > 1:
        user_input = choose_account(matches)
        if not user_input:
            exit()
            
    # รวมผลลัพธ์จากทุกคอลัมน์        
    filtered_df_list = []
    for col in account_cols:
        matches = df_gl[df_gl[col].astype(str).str.contains(user_input, case=False, na=False)]
        if not matches.empty:
            filtered_df_list.append(matches)

    if filtered_df_list:
        filtered_df = pd.concat(filtered_df_list, ignore_index=True).drop_duplicates()
    else:
        print("⚠️ ไม่พบข้อมูลที่ตรงกับบัญชีที่เลือก กรุณาลองใหม่!")
        continue
    break

# บันทึกข้อมูลที่กรองแล้ว
print(f"⌛ กำลังค้นหารายการ...")
with pd.ExcelWriter(output_filename, engine="openpyxl", mode="a") as writer:
    filtered_df.to_excel(writer, sheet_name=user_input, index=False)

print(f"💾 บันทึกรายการที่เลือกแล้ว")

# ทำการคำนวณและวิเคราะห์ข้อมูลตามที่ระบุในส่วนล่าง
df_selected = pd.read_excel(output_filename, sheet_name=user_input)

print("⌛ กำลังประมวลผล...")

# ตรวจสอบว่ามีคอลัมน์ 'เดบิต' หรือ 'Dr.' หรือไม่
debit_col = next((col for col in df_selected.columns if "เดบิต" in col.lower() or "dr." in col.lower()), None)
if not debit_col:
    print("❌ ไม่พบคอลัมน์เดบิตในชีทที่เลือก!")
    exit()

# ดึงข้อมูลจากคอลัมน์เดบิตและแปลงเป็นตัวเลข
valid_data = df_selected[debit_col].dropna().astype(float)

# ดึงตัวเลขหลักแรก
first_digits = valid_data.apply(lambda x: extract_first_digit(int(x)))

# คำนวณความถี่ของเลขหลักแรก
first_digit_counts = first_digits.value_counts().sort_index()
total_count = first_digit_counts.sum()

# Benford's Law ค่าความน่าจะเป็น
benford_probabilities = {d: np.log10(1 + 1/d) for d in range(1, 10)}
benford_predicted = {d: benford_probabilities[d] * total_count for d in benford_probabilities}

# 🔹 สร้างชีท Benford
wb = openpyxl.load_workbook(output_filename)
if "Benford" in wb.sheetnames:
    ws_benford = wb["Benford's"]
    wb.remove(ws_benford)
wb.create_sheet("Benford's")
ws_benford = wb["Benford's"]

# เพิ่มข้อมูล "รหัสบัญชี" และ "ชื่อบัญชี" จากข้อมูลที่กรองแล้ว
# ตรวจสอบว่า user_input เป็นรหัสบัญชีหรือชื่อบัญชี
if user_input.isdigit():  # หากผู้ใช้กรอกรหัสบัญชี
    account_no = user_input
    # ค้นหาและดึงชื่อบัญชีที่ตรงกับรหัสบัญชี
    account_name = filtered_df[filtered_df[account_cols[0]].astype(str) == account_no].iloc[0][account_cols[1]]
else:  # หากผู้ใช้กรอกชื่อบัญชี
    account_name = user_input
    # ค้นหาและดึงรหัสบัญชีที่ตรงกับชื่อบัญชี
    account_no = filtered_df[filtered_df[account_cols[1]].astype(str).str.contains(account_name, case=False, na=False)].iloc[0][account_cols[0]]

# เพิ่มข้อมูลรหัสบัญชีและชื่อบัญชีลงในชีท "Anomaly"
ws_benford.append([account_no, account_name])

# สีพื้นหลังและฟอนต์
header_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
header_font = openpyxl.styles.Font(bold=True, size=13)

# ทำให้แถวหัวมีสีพื้นหลังฟ้าและฟอนต์หนาขนาด 16
ws_benford.cell(row=1, column=1).fill = header_fill
ws_benford.cell(row=1, column=1).font = header_font
ws_benford.cell(row=1, column=2).fill = header_fill
ws_benford.cell(row=1, column=2).font = header_font
ws_benford.cell(row=1, column=3).fill = header_fill
ws_benford.cell(row=1, column=4).fill = header_fill
ws_benford.cell(row=1, column=5).fill = header_fill
ws_benford.cell(row=1, column=6).fill = header_fill
ws_benford.cell(row=1, column=7).fill = header_fill
ws_benford.cell(row=1, column=8).fill = header_fill

ws_benford.append(["", ""])

ws_benford.append(["DATA", "FIRST DIGIT"])

# ตั้งค่าสีพื้นหลังให้กับหัวข้อ
yel_fill = PatternFill(start_color="FFFF57", end_color="FFFF57", fill_type="solid")  # สีเขียวอ่อน
green_fill = PatternFill(start_color="A0EA68", end_color="A0EA68", fill_type="solid")  # สีเขียวอ่อน
head_font = openpyxl.styles.Font(bold=True, size=11)

# ทำให้แถวหัวมีสีพื้นหลังฟ้าและฟอนต์หนาขนาด 16
ws_benford.cell(row=3, column=1).fill = yel_fill
ws_benford.cell(row=3, column=2).fill = yel_fill
ws_benford.cell(row=3, column=1).font = head_font
ws_benford.cell(row=3, column=2).font = head_font

# ใส่ข้อมูล DATA และ FIRST DIGIT
for data, digit in zip(valid_data, first_digits):
    ws_benford.append([data, digit])

#  เว้น 1 คอลัมน์แล้วสร้างตาราง Benford's Law
start_row = 3  # เริ่มต้นจากแถวที่ 3
start_col = 4  # เริ่มต้นจากคอลัมน์ D (คอลัมน์ที่ 4)

# ค่าทฤษฎี Benford's Law สำหรับตัวเลขแรก 1-9
benford_predicted = {digit: math.log10(1 + 1 / digit) * 100 for digit in range(1, 10)}

# สมมติว่า 'total_count' และ 'first_digit_counts' ถูกกำหนดแล้ว
total_data_points = total_count

# กำหนดฟอนต์หนา
bold_font = Font(bold=True)

# เพิ่มหัวตารางในเซลล์ D4
ws_benford.cell(row=start_row, column=start_col, value="First Digit").fill = green_fill
ws_benford.cell(row=start_row, column=start_col, value="First Digit").font = bold_font

ws_benford.cell(row=start_row, column=start_col + 1, value="Occurrence of Digit").fill = green_fill
ws_benford.cell(row=start_row, column=start_col + 1, value="Occurrence of Digit").font = bold_font

ws_benford.cell(row=start_row, column=start_col + 2, value="Frequency of First Digit (%)").fill = green_fill
ws_benford.cell(row=start_row, column=start_col + 2, value="Frequency of First Digit (%)").font = bold_font

ws_benford.cell(row=start_row, column=start_col + 3, value="Predicted By Benford (%)").fill = green_fill
ws_benford.cell(row=start_row, column=start_col + 3, value="Predicted By Benford (%)").font = bold_font

# สร้างเส้นขอบ
border = Border(
    top=Side(border_style="thin", color="000000"),
    bottom=Side(border_style="thin", color="000000"),
    left=Side(border_style="thin", color="000000"),
    right=Side(border_style="thin", color="000000")
)

# เพิ่มเส้นขอบให้แถวนี้
for col in range(start_col, start_col + 4):
    ws_benford.cell(row=start_row, column=col).border = border

# คำนวณผลรวมของ "Occurrence", "Frequency", และ "Predicted"
total_occurrence = 0
total_frequency = 0
total_predicted = 0

pink_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")  # สีชมพูเข้ม
red_font = Font(color="FF0000", bold=True)

# เพิ่มข้อมูลตามลำดับในแถวถัดไป
for digit in range(1, 10):
    occurrence = first_digit_counts.get(digit, 0)
    predicted = benford_predicted[digit]  # ค่าที่คาดการณ์จากทฤษฎี Benford's Law
    frequency = (occurrence / total_data_points) * 100 if total_data_points else 0

    # คำนวณผลรวม
    total_occurrence += occurrence
    total_frequency += frequency
    total_predicted += predicted

    start_row += 1
    # สร้างเซลล์สำหรับ First Digit
    first_digit_cell = ws_benford.cell(row=start_row, column=start_col, value=digit)
    first_digit_cell.border = border

    # สร้างเซลล์สำหรับ Occurrence
    occurrence_cell = ws_benford.cell(row=start_row, column=start_col + 1, value=occurrence)
    occurrence_cell.border = border

    # ถ้า Frequency สูงกว่าค่า Predicted ให้เปลี่ยนสี **เฉพาะ First Digit และ Occurrence**
    if frequency > predicted:
        first_digit_cell.fill = pink_fill  # เปลี่ยนสีพื้นของ First Digit เป็นชมพูเข้ม
        occurrence_cell.fill = pink_fill  # เปลี่ยนสีพื้นของ Occurrence เป็นชมพูเข้ม

    # สร้างเซลล์สำหรับ Frequency และ Predicted 
    freq_cell = ws_benford.cell(row=start_row, column=start_col + 2, value=frequency)
    freq_cell.number_format = '0.00'
    freq_cell.border = border

    pred_cell = ws_benford.cell(row=start_row, column=start_col + 3, value=predicted)
    pred_cell.number_format = '0.00'
    pred_cell.border = border

    # ถ้า Frequency สูงกว่าค่า Predicted ให้เปลี่ยนสีเป็นชมพูเข้ม
    if frequency > predicted:
        freq_cell.fill = pink_fill  # เปลี่ยนสีพื้นเป็นชมพูเข้ม
        freq_cell.font = red_font  # เปลี่ยนฟอนต์เป็นสีแดงและหนา
        
# ทำให้ Total Frequency รวมเป็น 100%
if total_frequency != 100 and total_frequency > 0:
    adjustment_factor = 100 / total_frequency
    total_frequency = 0  # รีเซ็ตเพื่อคำนวณใหม่หลังการปรับค่า

    for digit in range(1, 10):
        row_idx = start_row - (9 - digit)  # คำนวณตำแหน่งแถวที่ต้องแก้ไข
        old_frequency = ws_benford.cell(row=row_idx, column=start_col + 2).value
        new_frequency = old_frequency * adjustment_factor  # ปรับค่าใหม่ให้รวมกันได้ 100%
        total_frequency += new_frequency  # รวมค่าที่ปรับแล้ว

        # อัปเดตค่า Frequency ใหม่ใน Excel
        ws_benford.cell(row=row_idx, column=start_col + 2, value=new_frequency).number_format = '0.00'

# เพิ่ม "Total Number Of Data Points" ลงในแถวถัดไป
start_row += 1
ws_benford.cell(row=start_row, column=start_col, value="Total Number Of Data Points").fill = header_fill
ws_benford.cell(row=start_row, column=start_col, value="Total Number Of Data Points").border = border
ws_benford.cell(row=start_row, column=start_col + 1, value=total_occurrence).border = border
ws_benford.cell(row=start_row, column=start_col + 2, value=total_frequency).number_format = '0.00'
ws_benford.cell(row=start_row, column=start_col + 2).border = border
ws_benford.cell(row=start_row, column=start_col + 3, value=total_predicted).number_format = '0.00'
ws_benford.cell(row=start_row, column=start_col + 3).border = border

# สร้างเส้นขอบ
border = Border(
    top=Side(border_style="thin", color="000000"),
    bottom=Side(border_style="thin", color="000000"),
    left=Side(border_style="thin", color="000000"),
    right=Side(border_style="thin", color="000000")
)

gray_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

# เพิ่มเส้นขอบให้แถวนี้
for col in range(start_col, start_col + 4):
    ws_benford.cell(row=start_row, column=col).border = border
    cell = ws_benford.cell(row=start_row, column=col)
    cell.fill = gray_fill

# อ้างอิงไปยังชีท "Anomaly"
ws = wb["Benford's"]

# ข้อมูลจากคอลัมป์ F2-F12 (Frequency of First Digit) และ G3-G12 (Predicted by Benford)
frequency_data = Reference(ws, min_col=6, min_row=3, max_row=12)  # คอลัมป์ F
benford_data = Reference(ws, min_col=7, min_row=3, max_row=12)  # คอลัมป์ G

# สร้างกราฟเส้น
chart = LineChart()
chart.add_data(frequency_data, titles_from_data=True)  # เพิ่มข้อมูลจริง
chart.add_data(benford_data, titles_from_data=True)  # เพิ่มข้อมูลที่คำนวณจาก Benford
chart.title = "Data Deviation from Benford"
chart.style = 3  # เลือกสไตล์ของกราฟ
chart.x_axis.title = "First Digit"
chart.y_axis.title = "Frequency (%)"
# ปรับขนาดภายนอก
chart.width = 16  # กว้าง
chart.height = 8  # สูง

chart.plot_area.left = 5  # เพิ่มพื้นที่ด้านซ้าย
chart.plot_area.right = 2  # เพิ่มพื้นที่ด้านขวา
chart.plot_area.top = 4    # เพิ่มพื้นที่ด้านบน
chart.plot_area.bottom = 4  # เพิ่มพื้นที่ด้านล่าง

chart.legend.position = 'b'  # ตั้งตำแหน่งคำอธิบายด้านล่าง
chart.legend.include_in_layout = False  # ปรับให้คำอธิบายไม่ซ้อนกับพื้นที่ของกราฟ
chart.legend.width = 10  # ปรับขนาดคำอธิบายให้เล็กลง
chart.legend.height = 2  # ปรับความสูงของคำอธิบาย

# เปิดการแสดง Gridlines
chart.x_axis.majorGridlines   # เปิดการแสดงกริดของแกน X
chart.y_axis.majorGridlines   # เปิดการแสดงกริดของแกน Y

# ตรวจสอบการแสดงตัวเลขบนแกน
chart.x_axis.delete = False
chart.y_axis.delete = False
chart.x_axis.tickLblPos = "nextTo"
chart.y_axis.tickLblPos = "nextTo"

# ตั้งค่า Tick Mark และ Major Unit
chart.x_axis.majorTickMark = "cross"
chart.y_axis.majorTickMark = "cross"
chart.x_axis.majorUnit = 1  # ให้แกน X แสดงค่าทีละ 1 (1,2,3,...9)
chart.y_axis.majorUnit = 10  # ให้แกน Y เพิ่มทีละ 10
chart.y_axis.minorUnit = 5   # ให้แกน Y มีค่าทศนิยมย่อยทีละ 5
chart.y_axis.majorTickMark = 'in'  # ให้ติ๊กบนแกน Y อยู่ภายในกราฟ

# ตั้งค่ารูปแบบตัวเลข
chart.x_axis.number_format = '0'
chart.y_axis.number_format = '0'

# ปรับตำแหน่งแกนให้แน่ใจว่ากราฟไม่ถูกบัง
chart.x_axis.overlap = 0
chart.y_axis.overlap = 0

# === ปรับลักษณะเส้นกราฟให้เหมือนตัวอย่าง ===

# ข้อมูลจริง (เส้นสีแดง)
series1 = chart.series[0]
series1.graphicalProperties.line.solidFill = "E60000"  # สีแดง
series1.graphicalProperties.line.width = 20000  # ความหนาเส้น
series1.marker.symbol = "none"  # จุดวงกลม

# ค่าคาดการณ์ Benford (เส้นสีน้ำเงิน)
series2 = chart.series[1]
series2.graphicalProperties.line.solidFill = "0070C0"  # สีน้ำเงิน
series2.graphicalProperties.line.width = 20000  # บางกว่าข้อมูลจริง
series2.marker.symbol = "none"  # จุดวงกลม

# กำหนดตำแหน่งของกราฟในชีท
ws.add_chart(chart, "D16")  # วางกราฟที่ตำแหน่ง D16

# 🔹 ฟังก์ชันปรับขนาดคอลัมน์ให้พอดีกับเนื้อหา และสามารถเลือกการจัดตำแหน่งได้
def autofit_columns(ws, alignment_dict, header_alignment="center"):
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter  # หาตัวอักษรคอลัมน์ เช่น A, B, C
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2  # ปรับระยะขอบเพิ่ม
        ws.column_dimensions[col_letter].width = adjusted_width

        for cell in col:
            # กำหนดการจัดตำแหน่งสำหรับส่วนหัว (header)
            if cell.row == 3:  # ส่วนหัวในแถวที่ 3
                header_alignment_obj = Alignment(horizontal=header_alignment)
                cell.alignment = header_alignment_obj
            else:
                # กำหนดการจัดตำแหน่งตาม dictionary สำหรับ data cells
                alignment_type = alignment_dict.get(col_letter, "center")  # ค่า default เป็น "center"
                
                if alignment_type == "left":
                    alignment = Alignment(horizontal="left")
                elif alignment_type == "right":
                    alignment = Alignment(horizontal="right")
                elif alignment_type == "center":
                    alignment = Alignment(horizontal="center")
                else:
                    alignment = Alignment(horizontal="center")

                # ตั้งค่าการจัดตำแหน่งให้กับทุกเซลล์ในคอลัมน์ (ยกเว้นส่วนหัว)
                if cell.row != 3:  # สำหรับแถวที่ไม่ใช่ส่วนหัว
                    cell.alignment = alignment

# 🔹 ตั้งการจัดตำแหน่งแต่ละคอลัมน์
alignment_dict = {
    "A": "right",   # คอลัมน์ A จัดชิดขวา
    "B": "center",  # คอลัมน์ B จัดกึ่งกลาง
    "C": "center",  # คอลัมน์ C จัดกึ่งกลาง
    "D": "center",  # คอลัมน์ D จัดกึ่งกลาง
    "E": "center",  # คอลัมน์ E จัดกึ่งกลาง
    "F": "right",   # คอลัมน์ F จัดชิดขวา
    "G": "right",   # คอลัมน์ G จัดชิดขวา
}

# ปรับขนาดคอลัมน์อัตโนมัติสำหรับชีท `Anomaly` และเลือกการจัดตำแหน่งตามที่กำหนด
autofit_columns(ws_benford, alignment_dict, header_alignment="center")  # ตั้งการจัดตำแหน่งส่วนหัวเป็น "center"

# ตรวจสอบว่ามีชีต "Analyze" หรือไม่ ถ้าไม่มีให้สร้างขึ้นมา
if "Analyze" not in wb.sheetnames:
    wb.create_sheet("Analyze")
    wb.save(output_filename)

# อ่านค่าตัวเลขหลักแรกที่มีสีชมพูจากชีต "Anomaly"
ws_benford = wb["Benford's"]
highlighted_digits = set()

# ค้นหาคอลัมน์ "First Digit", "Frequency" และ "Predicted"
first_digit_col = None
freq_col = None
pred_col = None

for col in ws_benford.iter_cols(min_row=3, max_row=3):
    header = col[0].value
    if header == "First Digit":
        first_digit_col = col[0].column
    elif header == "Frequency of First Digit (%)":
        freq_col = col[0].column
    elif header == "Predicted By Benford (%)":
        pred_col = col[0].column

# ตรวจสอบว่าพบทุกคอลัมน์ที่ต้องใช้หรือไม่
if None in (first_digit_col, freq_col, pred_col):
    print("❌ ไม่พบคอลัมน์ที่ต้องใช้ ('First Digit', 'Frequency of First Digit (%)', 'Predicted By Benford (%)')")
else:
    # ลูปอ่านข้อมูลตั้งแต่แถวที่ 4 ลงไป
    for row in ws_benford.iter_rows(min_row=4, max_row=12):
        first_digit_cell = row[first_digit_col - 1]  # First Digit
        freq_cell = row[freq_col - 1]  # Frequency
        pred_cell = row[pred_col - 1]  # Predicted

        # ตรวจสอบว่า Frequency > Predicted
        if freq_cell.value and pred_cell.value and freq_cell.value > pred_cell.value:
            highlighted_digits.add(first_digit_cell.value)  # เก็บเลขที่ผิดปกติ

# แสดงผลเลขที่ถูกไฮไลท์ใน First Digit
print(f"⚠️ พบตัวเลขหลักแรกที่มีความผิดปกติ : {highlighted_digits}")

# ตรวจสอบว่า user_input ตรงกับชื่อชีตที่มีอยู่ในไฟล์หรือไม่
if user_input in wb.sheetnames:
    account_sheets = [user_input]  # ใช้เฉพาะชีตที่ผู้ใช้เลือก
else:
    print(f"❌ ไม่พบชีต {user_input} ในไฟล์ Excel!")
    exit()
    
filtered_data = []

for sheet in account_sheets:
    df_account = pd.read_excel(output_filename, sheet_name=sheet)

    #  ค้นหาคอลัมน์ "เดบิต" หรือ "Dr."
    debit_col = next((col for col in df_account.columns if "เดบิต" in col.lower() or "dr." in col.lower()), None)
    if not debit_col:
        print(f"⚠️ ไม่พบคอลัมน์ 'เดบิต' ในชีต {sheet}, ข้ามไป!")
        continue  # ถ้าไม่มีข้อมูลเดบิต ให้ข้ามไป

     #  เพิ่มคอลัมน์ "First Digit" ก่อนคอลัมน์ "เดบิต"
    df_account["First Digit"] = pd.to_numeric(df_account[debit_col], errors='coerce').dropna().astype(str).str[0]
    df_account["First Digit"] = pd.to_numeric(df_account["First Digit"], errors='coerce').astype("Int64")

     # กรองเฉพาะรายการที่เลขหลักแรกตรงกับที่มีความผิดปกติ (จาก highlighted_digits)
    if highlighted_digits:
        df_filtered = df_account[df_account["First Digit"].isin(highlighted_digits)].copy()

        # ตรวจสอบว่ามีข้อมูลที่กรองได้หรือไม่
        if not df_filtered.empty:
            filtered_data.append(df_filtered)
            print(f"⚠️ พบ {len(df_filtered)} รายการที่ผิดปกติ !")
        else:
            print("⚠️ ไม่พบข้อมูลที่ตรงกับเลขหลักแรกที่ผิดปกติ !")

    else:
        print("❌ ไม่มีตัวเลขหลักแรกที่ผิดปกติ !")

#  โหลดข้อมูลที่ต้องบันทึก
df_result = pd.concat(filtered_data, ignore_index=True)

# บันทึกลงชีต "Analyze"
try:
    with pd.ExcelWriter(output_filename, engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
        df_result.to_excel(writer, sheet_name="Analyze", index=False)
    print("💾 บันทึกรายการผิดปกติเเล้ว")
except Exception as e:
    print(f"❌ เกิดข้อผิดพลาดขณะบันทึก: {e}")

print(f"📂 ดำเนินการเสร็จสิ้น สามารถตรวจสอบไฟล์ได้ที่ : {output_filename}")














